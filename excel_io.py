import os
import logging
import openpyxl
from copy import copy
from collections import defaultdict
from openpyxl.styles import Font, colors
import datetime
# from openpyxl.styles import Hyperlink

def get_history_aid(excel_path):
    aid_dict = defaultdict(list)
    if not os.path.isfile(excel_path):
        logging.warning(f'History excel not exist: {excel_path}')
        return aid_dict
    
    # 遍历 excel 获取每个account 的历史history
    # 打开Excel文件
    workbook = openpyxl.load_workbook(filename=excel_path)
    # 选择工作表
    sheet = workbook['Sheet']
    # 获取表头
    header = [cell.value for cell in sheet[1]]
    # 获取公众号和文章ID列的索引
    gzh_index = header.index('公众号')
    article_id_index = header.index('文章ID')
    # 构建字典
    data_dict = aid_dict
    for row in sheet.iter_rows(min_row=2, values_only=True):
        gzh_name = row[gzh_index]
        article_id = row[article_id_index]
        if gzh_name not in data_dict:
            data_dict[gzh_name] = [article_id]
        else:
            data_dict[gzh_name].append(article_id)
    return data_dict



def write_to_excel(file_path, data, sheet_name = 'Sheet'):
    """将数据写入Excel文件"""
    # 检查文件是否存在，如果不存在则创建一个新的Excel文件
    if not os.path.isfile(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    # 加载Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 检查sheet是否存在，如果不存在则创建一个新的sheet
    if sheet_name not in workbook.sheetnames:
        new_sheet = workbook.create_sheet(sheet_name)
    else:
        new_sheet = workbook[sheet_name]

    # 获取表头
    headers = []
    if new_sheet.max_row == 1:
        headers = list(data[0].keys())
        for col, header in enumerate(headers, start=1):
            new_sheet.cell(row=1, column=col, value=header)
        # 调整初始化列宽
        for col in new_sheet.columns:
            column = col[0].column_letter  # 获取列字母
            new_sheet.column_dimensions[column].width = 20
    else:
        for col in range(1, new_sheet.max_column + 1):
            headers.append(new_sheet.cell(row=1, column=col).value)

    num_existing_rows = new_sheet.max_row
    # 获取“发布时间”列的数据
    time_list = []
    for i in range(2, new_sheet.max_row + 1):
        time_value = new_sheet.cell(row=i, column=headers.index('发布时间') + 1).value
        if isinstance(time_value, datetime.datetime):
            time_list.append(time_value)
        else:
            time_list.append(datetime.datetime.strptime(time_value, '%Y-%m-%d %H:%M:%S'))

    # 插入新数据
    for item in data:
        row_to_insert = len(time_list) + 2
        for i in range(2, len(time_list) + 2):
            if datetime.datetime.strptime(item['发布时间'], '%Y-%m-%d %H:%M:%S') > time_list[i - 2]:
                row_to_insert = i
                break
        time_list.insert(row_to_insert - 2, datetime.datetime.strptime(item['发布时间'], '%Y-%m-%d %H:%M:%S'))
        new_sheet.insert_rows(row_to_insert)


        for col, value in item.items():
            col_index = headers.index(col) + 1
            new_sheet.cell(row=row_to_insert, column=col_index, value=value)
            # if col == '文章链接':
            #     # 如果该列是“文章链接”，则将单元格设置为超链接
            #     cell = new_sheet.cell(row=row_to_insert, column=col_index, value=value, hyperlink=value)
            #     cell.font = Font(color=colors.BLUE, underline='single')
            #     # Copy the hyperlink from the row above
            #     source_cell = worksheet.cell(row=row_idx-1, column=col_idx)
            #     if source_cell.hyperlink is not None:
            #         cell._hyperlink = copy(source_cell.hyperlink)
            # else:
            #     new_sheet.cell(row=row_to_insert, column=col_index, value=value)
    
    # 保存Excel文件
    workbook.save(file_path)



# def read_excel(file_path, sheet_name):
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook[sheet_name]
#     rows = list(sheet.rows)
#     headers = [cell.value for cell in rows[0]]

#     result = []
#     for row in rows[1:]:
#         item = {}
#         for col, cell in enumerate(row):
#             item[headers[col]] = cell.value
#         result.append(item)
#     return result

# def read_excel(file_path):
#     """_summary_

#     Args:
#         file_path (_type_): _description_

#     Returns:
#         Dict: [sheet_name]:[sheet_data]
#     """
#     workbook = openpyxl.load_workbook(file_path)
#     result = []
#     for sheet in workbook:
#         rows = list(sheet.rows)
#         headers = [cell.value for cell in rows[0]]
#         sheet_data = []
#         for row in rows[1:]:
#             item = {}
#             for col, cell in enumerate(row):
#                 item[headers[col]] = cell.value
#             sheet_data.append(item)
#         result.append({sheet.title: sheet_data})
#     return result


def read_all_columns(file_path, column_name):
    workbook = openpyxl.load_workbook(file_path)
    # 创建一个字典，用于存储每个sheet的数据
    data_dict = defaultdict(list)

    # 遍历所有sheet
    for sheet in workbook.worksheets:
        # 检查该sheet是否包含指定的列名
        col_index = 0
        for col in sheet.iter_cols(min_row=1, max_row=1, values_only=True):
            col_index += 1
            if column_name in col:
                break
        if sheet.cell(row=1, column=col_index).value != column_name:
            # 如果不包含，将该sheet的值设置为空列表
            data_dict[sheet.title] = []
            continue

        # 获取指定列的数据
        col_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            col_data.append(row[col_index - 1])

        # 将数据添加到字典中
        data_dict[sheet.title] = col_data

    # 返回包含所有sheet数据的字典
    return data_dict




if __name__ == '__main__':
    #search_aid('/Users/wangzepeng/Downloads/test.xlsx')
    data = [
    {"公众号": "A22A1", "发布时间": "2023-03-09 11:28:39", "文章标题": "baidu", '文章链接':'http://www.baidu.com', '文章ID':'12'},
    {"公众号": "A22A1", "发布时间": "2023-04-09 11:28:39", "文章标题": "google", '文章链接':'http://www.google.com', '文章ID':'12'},
    {"公众号": "A22A1", "发布时间": "2023-04-19 11:28:39", "文章标题": "ali", '文章链接':'http://www.alibaba.com', '文章ID':'12'},
    ]
    write_to_excel('./test.xlsx', data)
    data2 = [
    {"公众号": "A22A1", "发布时间": "2023-03-09 11:30:39", "文章标题": "shuiyuan", '文章链接':'http://www.shuiyuan.com', '文章ID':'12'},
    {"公众号": "A22A1", "发布时间": "2023-04-09 12:28:39", "文章标题": "weixin", '文章链接':'http://www.wechat.com', '文章ID':'12'},    ]
    write_to_excel('./test.xlsx', data2)
    get_history_aid(excel_path)
    # write_to_excel("example.xlsx", "Sheet", data)
    # write_to_excel("example.xlsx", "Sheet1", data)
    #read_all_columns("./wechat_report.xlsx", 'aid')
    